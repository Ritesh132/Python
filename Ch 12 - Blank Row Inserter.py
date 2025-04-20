
import sys
import os
import openpyxl


row_start = int(sys.argv[1])
row_length = int(sys.argv[2])
filename = sys.argv[3]
print('Opening workbook...')
os.chdir(r'C:\Users\rites\Desktop\Automate The Boring Stuff Files\automate_online-materials')
wb = openpyxl.load_workbook(filename)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

print('Adding blank rows...')
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        if row < row_start:
            new_sheet.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
        else:
            new_sheet.cell(row=row+row_length, column=column).value = sheet.cell(row=row, column=column).value


os.chdir(r'c:\users\rites\desktop')
new_wb.save('blank_row.xlsx')
print('Done.')
