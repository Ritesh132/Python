import sys
import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

os.chdir(r'c:\users\tomal\desktop')

n_table = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active

for r in range(2, n_table + 2):
    sheet['A' + str(r)] = r - 1
    col_let = get_column_letter(r)
    sheet[col_let + str(1)] = r - 1
    # Set font to bold
    sheet['A' + str(r)].font = Font(bold=True)
    sheet[col_let + str(1)].font = Font(bold=True)

for col in range(2, n_table + 2):
    for row in range(2, n_table + 2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col - 1) * (row - 1)

wb.save('Multiplication Table.xlsx')
